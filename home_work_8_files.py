import csv
import os
import shutil
import sys
from openpyxl import load_workbook
import zipfile
import PyPDF2

# разархивируем файлы
zip_1 = zipfile.ZipFile('resources/sample.zip')
print(zip_1.namelist())
zip_1.extractall('tmp')

# архивируем обратно
zip_2 = zipfile.ZipFile('test.zip', 'w')
for root, dirs, files in os.walk(r'./resources'):
    for file in files:
        zip_2.write(os.path.join(root, file))
    zip_2.close()


# читаем файлы
def read_the_files(file_end):
    arhive = 'resources/sample.zip'
    zip_file = zipfile.ZipFile(arhive)
    path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'tmp')
    for file in zip_file.infolist():
        if file.filename.endswith('.pdf') and file_end == '.pdf':
            print('PDF')
            pdf_file = zip_file.extract(file.filename, 'tmp')
            plread = PyPDF2.PdfFileReader(pdf_file)
            getpage2 = plread.getPage(2)
            text2 = getpage2.extractText()
            assert 'Simple PDF File 2' in text2
            shutil.rmtree(path)
            sys.exit()
    if file.filename.endswith('.csv') and file_end == '.csv':
        print('CSV')
        csv_file = zip_file.extract(file.filename, 'tmp')
        with open(csv_file) as csvfile:
            csvfile = csv.reader(csvfile)
            for strings in csvfile:
                if 'John' in strings:
                    assert True
                    break
            else:
                assert False
        shutil.rmtree(path)
        sys.exit()
    if file.filename.endswith('.xlsx') and file_end == '.xlsx':
        print('XLSX')
        xlsx_file = zip_file.extract(file.filename, 'tmp')
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active
        list_fom_cells = []
        for row in range(sheet.max_row):
            for column in range(sheet.max_column):
                value_str = str(sheet.cell(row=(row + 1), column=(column + 1)).value)
                list_fom_cells.append(value_str)
        if '1' in list_fom_cells:
            assert True

        else:
            assert False
        shutil.rmtree(path)
        sys.exit()


read_the_files('pdf')
