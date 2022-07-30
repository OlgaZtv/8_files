import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook


# читаем файлы
def test_read_pdf():
    reader = PdfReader('resources/sample.pdf')
    page = reader.pages[0]
    text = page.extract_text()
    assert 'A Simple PDF File' in text


def test_read_xlsx():
    workbook = load_workbook('resources/file_example.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=2, column=2).value
    assert 'Dulce' == name


def test_read_csv_():
    with open('resources/example.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'John' in str(headers)
